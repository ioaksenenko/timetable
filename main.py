import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils.cell import get_column_letter
from argparse import ArgumentParser


def get_classrooms(wb):
    ws = wb['Расписание']
    res = {}
    weekday = None
    for i, row in enumerate(ws.rows):
        for j, col in enumerate(row):
            match = re.fullmatch(r'.+-(?P<classroom_number>\d+(-\d)?)(/(?P<alternative_number>\d+(-\d)?))?', str(col.value))
            if match:
                classroom_number = match.group('classroom_number')
                alternative_number = match.group('alternative_number')
                weekday = row[0].value or weekday
                lesson = row[1].value
                if weekday and lesson:
                    if weekday not in res:
                        res[weekday] = {}
                    if lesson not in res[weekday]:
                        res[weekday][lesson] = {
                            'numbers': [],
                            'address': []
                        }
                    res[weekday][lesson]['numbers'].append(classroom_number)
                    address = f'{get_column_letter(j + 1)}{i + 1}'
                    res[weekday][lesson]['address'].append(address)
                    if alternative_number:
                        res[weekday][lesson]['numbers'].append(alternative_number)
                        res[weekday][lesson]['address'].append(address)
    return res


def fill_workload(wb, classrooms):
    ws = wb['Загрузка кабинетов']
    weekday_index = None
    lesson_index = None
    weekday = None
    lesson = None
    for i, row in enumerate(ws.rows):
        classroom_number = str(row[0].value) if re.fullmatch(r'\d+(-\d+)?', str(row[0].value)) else None
        for j, col in enumerate(row):
            if classroom_number and not col.value:
                if not lesson_index and not weekday_index:
                    weekday_index = i - 1
                    lesson_index = i
                weekday = ws[weekday_index][j].value or weekday
                lesson = ws[lesson_index][j].value or lesson
                if weekday in classrooms and lesson in classrooms[weekday] and classroom_number in classrooms[weekday][
                    lesson]['numbers']:
                    ws[i + 1][j].value = '+'
                    ws[i + 1][j].font = Font(color="000000")
                    if classrooms[weekday][lesson]['numbers'].count(classroom_number) > 1:
                        ws[i + 1][j].fill = PatternFill(fgColor="FFC7CE", fill_type="solid")
                        col_numbers = ",".join([classrooms[weekday][lesson]['address'][k] for k, cl_n in enumerate(
                            classrooms[weekday][lesson]['numbers']
                        ) if cl_n == classroom_number])
                        ws[i + 1][j].comment = Comment(f'Адреса ячеек: {col_numbers}', 'Automatic comment')


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('-in', '--input', dest='input_file_path', help='Input file path', required=True)
    parser.add_argument('-out', '--output', dest='output_file_path', default='output.xlsx', help='Output file path')
    args = parser.parse_args()

    wb = load_workbook(args.input_file_path)
    classrooms = get_classrooms(wb)
    fill_workload(wb, classrooms)
    wb.save(args.output_file_path)
